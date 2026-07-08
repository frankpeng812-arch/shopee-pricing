#!/usr/bin/env python3
"""
将 Shopee 官方「跨境物流成本（藏价）」Excel 解析为 shipping_rules.json。

用法:
  python parse_shipping_excel.py "跨境物流成本（藏价）计算工具 - 20260604起生效.xlsx"
  python parse_shipping_excel.py input.xlsx -o shipping_rules.json --verbose
  python parse_shipping_excel.py --generate-sample sample.xlsx   # 生成测试用样例表
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# 站点映射：Excel 文本 -> (site_code, 中文名, 货币 ISO)
# ---------------------------------------------------------------------------
SITE_RULES: list[tuple[re.Pattern[str], tuple[str, str, str]]] = [
    (re.compile(r"新加坡|singapore|\bSGD\b|\bSG\b", re.I), ("SG", "新加坡", "SGD")),
    (re.compile(r"马来西亚|malaysia|\bMYR\b|\bMY\b", re.I), ("MY", "马来西亚", "MYR")),
    (re.compile(r"泰国|thailand|\bTHB\b|\bTH\b", re.I), ("TH", "泰国", "THB")),
    (re.compile(r"菲律宾|philippines|\bPHP\b|\bPH\b", re.I), ("PH", "菲律宾", "PHP")),
    (re.compile(r"越南|vietnam|\bVND\b|\bVN\b", re.I), ("VN", "越南", "VND")),
    (re.compile(r"台湾|taiwan|\bTWD\b|\bTW\b", re.I), ("TW", "台湾", "TWD")),
    (re.compile(r"巴西|brazil|\bBRL\b|\bBR\b", re.I), ("BR", "巴西", "BRL")),
    (re.compile(r"墨西哥|mexico|\bMXN\b|\bMX\b", re.I), ("MX", "墨西哥", "MXN")),
    (re.compile(r"哥伦比亚|colombia|\bCOP\b|\bCO\b", re.I), ("CO", "哥伦比亚", "COP")),
    (re.compile(r"智利|chile|\bCLP\b|\bCL\b", re.I), ("CL", "智利", "CLP")),
]

# ---------------------------------------------------------------------------
# 特殊/复合渠道：正则难以稳定解析时直接硬编码或跳过
# key 为渠道名子串（不区分大小写），value 为规则或 {"skip": True}
# ---------------------------------------------------------------------------
SPECIAL_CHANNELS: dict[str, dict[str, Any]] = {
    "sea shipping": {
        "base_weight": 2000,
        "base_price": 3.4,
        "step_weight": 1000,
        "step_price": 0.4,
        "note": "海运按大重量段计费，需人工核对官方表",
    },
    "海运": {
        "base_weight": 2000,
        "base_price": 3.4,
        "step_weight": 1000,
        "step_price": 0.4,
    },
    "economy sea": {"skip": True, "reason": "经济海运计费复杂，请手动维护"},
    "轻货": {"skip": True, "reason": "轻货渠道按体积/特殊段计费"},
}

# 表头列名模糊匹配
COLUMN_ALIASES: dict[str, list[str]] = {
    "site": ["站点", "国家", "市场", "site", "country"],
    "channel": ["物流方案", "物流渠道", "渠道", "delivery", "shipping method", "物流"],
    "zone": ["地区", "分区", "区域", "zone", "region"],
    "base_price": ["起重", "首重", "起重价格", "首重价格", "藏价起重"],
    "step_price": ["续重", "续重价格", "续费"],
}

SHEET_NAME_HINTS = [
    "物流成本价格详解",
    "跨境物流成本（藏价）计算工具",
    "跨境物流成本",
    "藏价",
    "价格详解",
]

SECTION_SKIP_KEYWORDS = [
    "默认价格表",
    "特货价格表",
    "重货价格表",
    "轻货价格表",
    "说明",
    "例子",
    "示例",
    "注：",
    "注:",
]

# 起重 / 续重 正则
BASE_PRICE_PATTERNS = [
    re.compile(
        r"[≤<=≤]\s*(\d+(?:\.\d+)?)\s*[gG克].*?[:：]\s*(\d+(?:\.\d+)?)",
        re.I,
    ),
    re.compile(
        r"(\d+(?:\.\d+)?)\s*[gG克]\s*(?:以内|以下|之内).*?[:：]?\s*(\d+(?:\.\d+)?)",
        re.I,
    ),
    re.compile(r"首重\s*(\d+(?:\.\d+)?)\s*[gG克].*?(\d+(?:\.\d+)?)", re.I),
    re.compile(r"(\d+(?:\.\d+)?)\s*[gG克].*?(\d+(?:\.\d+)?)\s*(?:SGD|MYR|THB|PHP|TWD|VND|CNY)?$", re.I),
]

STEP_PRICE_PATTERNS = [
    re.compile(
        r">\s*(\d+(?:\.\d+)?)\s*[gG克].*?[:：]\s*(\d+(?:\.\d+)?)\s*[/／]\s*(\d+(?:\.\d+)?)\s*[gG克]",
        re.I,
    ),
    re.compile(
        r"续重.*?(\d+(?:\.\d+)?)\s*[/／]\s*(\d+(?:\.\d+)?)\s*[gG克].*?[:：]?\s*(\d+(?:\.\d+)?)",
        re.I,
    ),
    re.compile(
        r"每\s*(\d+(?:\.\d+)?)\s*[gG克].*?[:：]?\s*(\d+(?:\.\d+)?)",
        re.I,
    ),
    re.compile(
        r"(\d+(?:\.\d+)?)\s*[/／]\s*(\d+(?:\.\d+)?)\s*[gG克].*?[:：]?\s*(\d+(?:\.\d+)?)",
        re.I,
    ),
]

SITE_BY_CODE: dict[str, tuple[str, str, str]] = {
    code: (code, name, currency) for _, (code, name, currency) in SITE_RULES
}

VERSION_FROM_FILENAME = re.compile(r"(20\d{6})")


def normalize_cell(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def map_site(raw: str) -> tuple[str, str, str] | None:
    text = normalize_cell(raw)
    if not text:
        return None
    for pattern, meta in SITE_RULES:
        if pattern.search(text):
            return meta
    return None


def extract_version(path: Path, sheet_names: list[str]) -> str:
    match = VERSION_FROM_FILENAME.search(path.stem)
    if match:
        return match.group(1)
    for name in sheet_names:
        match = VERSION_FROM_FILENAME.search(name)
        if match:
            return match.group(1)
    return date.today().strftime("%Y%m%d")


def pick_sheet(wb: openpyxl.Workbook, preferred: str | None) -> str:
    if preferred:
        if preferred in wb.sheetnames:
            return preferred
        raise ValueError(f"指定 Sheet 不存在: {preferred}，可选: {wb.sheetnames}")

    for hint in SHEET_NAME_HINTS:
        for name in wb.sheetnames:
            if hint in name:
                return name

    for name in wb.sheetnames:
        lower = name.lower()
        if "详解" in name or "藏价" in name:
            return name

    raise ValueError(
        "未找到合适的 Sheet，请用 --sheet 指定。"
        f" 当前工作簿包含: {wb.sheetnames}"
    )


def row_looks_like_header(cells: list[str]) -> bool:
    joined = " ".join(cells).lower()
    has_site = any(alias in joined for alias in COLUMN_ALIASES["site"])
    has_channel = any(alias in joined for alias in COLUMN_ALIASES["channel"])
    has_price = any(
        alias in joined
        for alias in COLUMN_ALIASES["base_price"] + COLUMN_ALIASES["step_price"]
    )
    return has_site and has_channel and has_price


def detect_columns(header: list[str]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {key: None for key in COLUMN_ALIASES}
    for idx, cell in enumerate(header):
        lower = cell.lower()
        for field, aliases in COLUMN_ALIASES.items():
            if mapping[field] is not None:
                continue
            if any(alias.lower() in lower for alias in aliases):
                mapping[field] = idx
    return mapping


def should_skip_row(site: str, channel: str, base_text: str, step_text: str) -> bool:
    probe = f"{site} {channel} {base_text} {step_text}".strip()
    if not probe:
        return True
    if any(kw in probe for kw in SECTION_SKIP_KEYWORDS):
        return True
    if channel and channel.endswith("表"):
        return True
    return False


def match_special_channel(channel: str) -> dict[str, Any] | None:
    lower = channel.lower()
    for key, rule in SPECIAL_CHANNELS.items():
        if key.lower() in lower:
            return rule
    return None


def parse_base_price(text: str) -> tuple[float, float] | None:
    text = normalize_cell(text)
    if not text:
        return None
    for pattern in BASE_PRICE_PATTERNS:
        match = pattern.search(text)
        if match:
            return float(match.group(1)), float(match.group(2))
    return None


def parse_step_price(text: str, base_weight_hint: float | None = None) -> tuple[float, float] | None:
    text = normalize_cell(text)
    if not text:
        return None

    for pattern in STEP_PRICE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        groups = match.groups()
        if len(groups) == 3:
            # >40g: 0.15 / 10g  =>  step_weight=10, step_price=0.15
            threshold, price_or_step, step_or_price = (
                float(groups[0]),
                float(groups[1]),
                float(groups[2]),
            )
            if price_or_step < step_or_price:
                step_price, step_weight = price_or_step, step_or_price
            else:
                step_weight, step_price = price_or_step, step_or_price
            return step_weight, step_price
        if len(groups) == 2:
            # 每 10g: 0.15
            step_weight, step_price = float(groups[0]), float(groups[1])
            return step_weight, step_price

    if base_weight_hint is not None:
        plain = re.search(r"(\d+(?:\.\d+)?)", text)
        if plain:
            return 10.0, float(plain.group(1))

    return None


def build_channel_key(channel: str, zone: str) -> str:
    channel = normalize_cell(channel)
    zone = normalize_cell(zone)
    if zone and zone.lower() not in {"nan", "-", "—", "全部", "all"}:
        return f"{channel} ({zone})"
    return channel


def parse_pricing_row(
    site_raw: str,
    channel_raw: str,
    zone_raw: str,
    base_raw: str,
    step_raw: str,
) -> tuple[str, dict[str, Any]] | None:
    site_meta = map_site(site_raw)
    if not site_meta:
        return None

    site_code, _, _ = site_meta
    channel = normalize_cell(channel_raw)
    if not channel:
        return None

    if should_skip_row(site_raw, channel, base_raw, step_raw):
        return None

    special = match_special_channel(channel)
    if special:
        if special.get("skip"):
            return None
        rule = {
            "base_weight": special["base_weight"],
            "base_price": special["base_price"],
            "step_weight": special["step_weight"],
            "step_price": special["step_price"],
        }
        if special.get("note"):
            rule["_note"] = special["note"]
        channel_key = build_channel_key(channel, zone_raw)
        return site_code, {"channel_key": channel_key, "rule": rule}

    base_parsed = parse_base_price(base_raw)
    if not base_parsed:
        return None

    base_weight, base_price = base_parsed
    step_parsed = parse_step_price(step_raw, base_weight)
    if not step_parsed:
        return None

    step_weight, step_price = step_parsed

    channel_key = build_channel_key(channel, zone_raw)
    rule = {
        "base_weight": base_weight,
        "base_price": base_price,
        "step_weight": step_weight,
        "step_price": step_price,
    }
    return site_code, {"channel_key": channel_key, "rule": rule}


def sheet_to_matrix(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    matrix: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append([normalize_cell(v) for v in row])
    return matrix


def forward_fill_columns(matrix: list[list[str]], col_indices: list[int]) -> None:
    last: dict[int, str] = {}
    for row in matrix:
        for col in col_indices:
            if col >= len(row):
                continue
            if row[col]:
                last[col] = row[col]
            elif col in last:
                row[col] = last[col]


def iter_data_blocks(matrix: list[list[str]]) -> list[tuple[list[str], list[list[str]]]]:
    """扫描整张 Sheet，支持多个「表头 + 数据区」块（默认/特货/重货等）。"""
    blocks: list[tuple[list[str], list[list[str]]]] = []
    i = 0
    n = len(matrix)

    while i < n:
        row = matrix[i]
        if row_looks_like_header(row):
            header = row
            i += 1
            data_rows: list[list[str]] = []
            while i < n:
                next_row = matrix[i]
                if row_looks_like_header(next_row):
                    break
                if any(next_row):
                    data_rows.append(next_row)
                i += 1
            blocks.append((header, data_rows))
            continue
        i += 1

    return blocks


def parse_workbook(
    path: Path,
    sheet_name: str | None = None,
    verbose: bool = False,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    target_sheet = pick_sheet(wb, sheet_name)
    ws = wb[target_sheet]
    matrix = sheet_to_matrix(ws)

    blocks = iter_data_blocks(matrix)
    if not blocks:
        raise ValueError(f"Sheet「{target_sheet}」中未识别到有效表头，请检查 Excel 格式。")

    data: dict[str, dict[str, Any]] = {}
    stats = {"parsed": 0, "skipped": 0, "blocks": len(blocks)}

    for block_idx, (header, rows) in enumerate(blocks):
        columns = detect_columns(header)
        required = ["site", "channel", "base_price", "step_price"]
        missing = [k for k in required if columns[k] is None]
        if missing:
            if verbose:
                print(f"[warn] 块 #{block_idx + 1} 缺少列 {missing}，跳过", file=sys.stderr)
            continue

        fill_cols = [c for c in [columns["site"], columns["channel"]] if c is not None]
        forward_fill_columns(rows, fill_cols)

        zone_col = columns.get("zone")

        for row in rows:
            site_raw = row[columns["site"]] if columns["site"] is not None else ""
            channel_raw = row[columns["channel"]] if columns["channel"] is not None else ""
            zone_raw = row[zone_col] if zone_col is not None and zone_col < len(row) else ""
            base_raw = row[columns["base_price"]] if columns["base_price"] is not None else ""
            step_raw = row[columns["step_price"]] if columns["step_price"] is not None else ""

            parsed = parse_pricing_row(site_raw, channel_raw, zone_raw, base_raw, step_raw)
            if not parsed:
                stats["skipped"] += 1
                continue

            site_code, payload = parsed
            site_meta = map_site(site_raw) or SITE_BY_CODE.get(site_code)
            if not site_meta:
                stats["skipped"] += 1
                continue
            _, site_name, currency = site_meta

            if site_code not in data:
                data[site_code] = {
                    "name": site_name,
                    "currency": currency,
                    "channels": {},
                }

            channel_key = payload["channel_key"]
            rule = payload["rule"]
            data[site_code]["channels"][channel_key] = rule
            stats["parsed"] += 1

    if not data:
        raise ValueError("未能从 Excel 提取任何有效渠道规则，请检查表头与单元格格式。")

    version = extract_version(path, wb.sheetnames)
    result = {
        "version": version,
        "updated_at": date.today().isoformat(),
        "source_sheet": target_sheet,
        "stats": stats,
        "data": data,
    }
    return result


def write_json(payload: dict[str, Any], output_path: Path) -> None:
    cleaned_data: dict[str, Any] = {}
    for site_code, site in payload["data"].items():
        channels: dict[str, Any] = {}
        for name, rule in site["channels"].items():
            channels[name] = {
                k: v
                for k, v in rule.items()
                if not str(k).startswith("_")
            }
        cleaned_data[site_code] = {
            "name": site["name"],
            "currency": site["currency"],
            "channels": channels,
        }

    public_payload = {
        "version": payload["version"],
        "updated_at": payload["updated_at"],
        "data": cleaned_data,
    }
    output_path.write_text(
        json.dumps(public_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def generate_sample_excel(path: Path) -> None:
    """生成与官方排版类似的测试 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物流成本价格详解"

    ws.append(["默认价格表"])
    ws.append([])
    ws.append(["站点", "物流方案/渠道", "地区/分区", "起重价格", "续重价格"])
    rows = [
        ["新加坡 (SGD)", "Standard Doorstep Delivery", "Zone A", "≤40g: 1.26", ">40g: 0.15 / 10g"],
        ["", "Standard Doorstep Delivery", "Zone B", "≤40g: 1.30", ">40g: 0.16 / 10g"],
        ["马来西亚 (MYR)", "Standard Doorstep Delivery", "", "≤10g: 0.15", ">10g: 0.15 / 10g"],
        ["泰国 (THB)", "Standard Delivery", "Zone KV", "≤10g: 1", ">10g: 1 / 10g"],
        ["菲律宾 (PHP)", "Standard Delivery", "", "≤50g: 23", ">50g: 4.5 / 10g"],
        ["新加坡 (SGD)", "Doorstep Delivery (Sea Shipping)", "", "≤2000g: 3.4", ">2000g: 0.4 / 1000g"],
    ]
    for row in rows:
        ws.append(row)

    wb.save(path)
    print(f"已生成样例 Excel: {path}")


def run_self_tests() -> None:
    assert parse_base_price("≤40g: 1.26") == (40.0, 1.26)
    assert parse_step_price(">40g: 0.15 / 10g", 40) == (10.0, 0.15)
    assert map_site("新加坡 (SGD)")[0] == "SG"
    assert map_site("马来西亚MYR")[0] == "MY"

    site_code, payload = parse_pricing_row(
        "新加坡 (SGD)",
        "Standard Doorstep Delivery",
        "Zone A",
        "≤40g: 1.26",
        ">40g: 0.15 / 10g",
    )
    assert site_code == "SG"
    assert payload["rule"]["base_weight"] == 40

    special = parse_pricing_row(
        "新加坡 (SGD)",
        "Doorstep Delivery (Sea Shipping)",
        "",
        "复杂格式",
        "无法解析",
    )
    assert special is not None
    assert special[1]["rule"]["base_weight"] == 2000

    print("self-tests passed")


def main() -> int:
    parser = argparse.ArgumentParser(description="解析 Shopee 官方藏价 Excel -> shipping_rules.json")
    parser.add_argument("input", nargs="?", help="官方 Excel 文件路径")
    parser.add_argument(
        "-o",
        "--output",
        default="shipping_rules.json",
        help="输出 JSON 路径（默认：脚本同级 shipping_rules.json）",
    )
    parser.add_argument("--sheet", help="指定 Sheet 名称")
    parser.add_argument("-v", "--verbose", action="store_true", help="输出解析统计")
    parser.add_argument("--generate-sample", metavar="PATH", help="生成测试用样例 Excel")
    parser.add_argument("--test", action="store_true", help="运行内置单元测试")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent

    if args.test:
        run_self_tests()
        return 0

    if args.generate_sample:
        out = Path(args.generate_sample)
        if not out.is_absolute():
            out = script_dir / out
        generate_sample_excel(out)
        return 0

    if not args.input:
        parser.error("请提供 input.xlsx，或使用 --generate-sample / --test")

    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = Path.cwd() / input_path
    if not input_path.exists():
        print(f"文件不存在: {input_path}", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = script_dir / output_path

    try:
        payload = parse_workbook(input_path, sheet_name=args.sheet, verbose=args.verbose)
        write_json(payload, output_path)
    except Exception as exc:
        print(f"解析失败: {exc}", file=sys.stderr)
        return 1

    print(f"✓ 已写入 {output_path}")
    if args.verbose:
        stats = payload.get("stats", {})
        print(
            f"  Sheet: {payload.get('source_sheet')} | "
            f"version: {payload['version']} | "
            f"站点: {len(payload['data'])} | "
            f"规则: {stats.get('parsed', 0)} | "
            f"跳过: {stats.get('skipped', 0)}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
